from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from fpdf import FPDF
import pandas as pd
import tempfile

app = Flask(__name__)
app.secret_key = 'your_secret_key'

MASTER_FILE = "master_data.xlsx"
USER_FILE = "users.xlsx"

# ------------------ Utility Functions ------------------

def load_data():
    if not os.path.exists(MASTER_FILE):
        wb = Workbook()
        wb.save(MASTER_FILE)
    wb = load_workbook(MASTER_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else []
    data = [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    return data

def load_users():
    if not os.path.exists(USER_FILE):
        return []
    wb = load_workbook(USER_FILE)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]

def authenticate_user(email, password):
    users = load_users()
    return next((user for user in users if user.get("Email") == email and user.get("Password") == password), None)

def filter_data(data, filters):
    filtered, seen_keys = [], set()
    for row in data:
        match = True
        for key, value in filters.items():
            if key == "warranty_expiry" and value == "1":
                expiry = row.get("Warranty End Date", "")
                try:
                    expiry_date = datetime.strptime(str(expiry), "%Y-%m-%d")
                    if expiry_date > datetime.today() + timedelta(days=30):
                        match = False
                except:
                    match = False
            elif key == "updated_after" and value:
                try:
                    updated_date = datetime.strptime(str(row.get("Last Updated Date", "")), "%Y-%m-%d")
                    if updated_date < datetime.strptime(value, "%Y-%m-%d"):
                        match = False
                except:
                    match = False
            elif key == "latest_only" and value == "1":
                unique_id = row.get("Serial Number") or row.get("Asset Name")
                if unique_id in seen_keys:
                    match = False
                seen_keys.add(unique_id)
            elif value and key in row and str(value).lower() not in str(row[key]).lower():
                match = False
        if match:
            filtered.append(row)
    return filtered

# ------------------ Routes ------------------

@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user' in session else 'login')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = authenticate_user(request.form['email'], request.form['password'])
        if user:
            session['user'] = user['Email']
            session['role'] = user.get('Role', 'user')
            return redirect(url_for('dashboard'))
        flash('Invalid credentials', 'danger')
    return render_template('login.html')

@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))

    data = load_data()
    filters = {
        "Asset Name": request.form.get("asset_name", ""),
        "Asset Type": request.form.get("asset_type", ""),
        "Model": request.form.get("model", ""),
        "Installed Date": request.form.get("installed_date", ""),
        "Working Condition": request.form.get("working_condition", ""),
        "Installation Status": request.form.get("installation_status", ""),
        "Location Installed": request.form.get("location", ""),
        "Vendor": request.form.get("vendor", ""),
        "warranty_expiry": request.form.get("warranty_expiry", ""),
        "updated_after": request.form.get("updated_after", ""),
        "latest_only": request.form.get("latest_only", "")
    } if request.method == 'POST' else {}

    session['filtered_data'] = filter_data(data, filters) if filters else data
    return render_template("dashboard.html", data=session['filtered_data'], is_admin=(session.get("role") == "admin"))

@app.route('/upload_asset', methods=['GET', 'POST'])
def upload_asset():
    if 'user' not in session or session.get('role') != 'admin':
        flash("Unauthorized: Only admins can upload assets.", "danger")
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        file = request.files.get('asset_file')
        if file and file.filename.endswith('.xlsx'):
            wb_uploaded = load_workbook(file)
            ws_uploaded = wb_uploaded.active

            if not os.path.exists(MASTER_FILE):
                wb_master = Workbook()
                ws_master = wb_master.active
                ws_master.append([cell.value for cell in ws_uploaded[1]])
            else:
                wb_master = load_workbook(MASTER_FILE)
                ws_master = wb_master.active

            for row in ws_uploaded.iter_rows(min_row=2, values_only=True):
                ws_master.append(list(row))

            wb_master.save(MASTER_FILE)
            flash("File uploaded and data appended successfully.", "success")
            return redirect(url_for('dashboard'))
        flash("Please upload a valid .xlsx file.", "warning")
        return redirect(url_for('upload_asset'))

    return render_template('upload_asset.html')

@app.route('/add_asset', methods=['POST'])
def add_asset():
    if 'user' not in session or session.get('role') != 'admin':
        flash("Unauthorized: Only admins can add assets.", "danger")
        return redirect(url_for('dashboard'))

    fields = [
        "Asset Name", "Asset Type", "Model", "Serial Number", "Installed Date",
        "Working Condition", "Installation Status", "Location Installed",
        "Warranty End Date", "Vendor"
    ]
    new_data = {field: request.form.get(field.replace(" ", "_").lower(), "") for field in fields}
    new_data["Last Updated Date"] = datetime.now().strftime("%Y-%m-%d")

    wb = load_workbook(MASTER_FILE) if os.path.exists(MASTER_FILE) else Workbook()
    ws = wb.active
    if ws.max_row < 1:
        ws.append(fields + ["Last Updated Date"])

    ws.append([new_data.get(field, "") for field in fields] + [new_data["Last Updated Date"]])
    wb.save(MASTER_FILE)

    flash("Asset added successfully", "success")
    return redirect(url_for('dashboard'))

@app.route('/export_pdf')
def export_pdf():
    if 'user' not in session:
        return redirect(url_for('login'))

    data = session.get('filtered_data', load_data())
    if not data:
        flash("No data available to export", "warning")
        return redirect(url_for('dashboard'))

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(0, 10, "KPL Asset Report", ln=True, align='C')
    pdf.ln(4)

    headers = list(data[0].keys())
    col_width = 270 / len(headers)

    pdf.set_font("Arial", 'B', 9)
    for header in headers:
        pdf.cell(col_width, 10, str(header), 1)
    pdf.ln()

    pdf.set_font("Arial", size=9)
    for row in data:
        for key in headers:
            pdf.cell(col_width, 10, str(row.get(key, ""))[:25], 1)
        pdf.ln()

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    pdf.output(tmp.name)
    return send_file(tmp.name, as_attachment=True, download_name="KPL_Asset_Report.pdf")

@app.route('/export/excel')
def export_excel():
    if 'user' not in session:
        return redirect(url_for('login'))

    data = session.get('filtered_data', load_data())
    if not data:
        flash("No data available to export", "warning")
        return redirect(url_for('dashboard'))

    df = pd.DataFrame(data)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    df.to_excel(tmp.name, index=False)
    return send_file(tmp.name, as_attachment=True, download_name="KPL_Asset_Report.xlsx")

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

if __name__ == '__main__':
    app.run(debug=True)
